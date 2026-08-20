"""Compile data/menu.xlsx into src/data/menu.json.

Schema v3 additions:
- Simple items gain optional subcategory / deal / boxPrice.
- Sheets 特色餐 (special) and 水果 (fruit) are first-class categories.
- 冰箱雜貨 maps to the grocery category (displayed as 冰箱雜貨).
- 加購規則 sheet becomes structured addon/upgrade/option/info rules.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "menu.xlsx"
OUT = ROOT / "src" / "data" / "menu.json"

SIZE_KEYS = ("hot_s", "hot_m", "iced_m", "xl")
SIMPLE_SHEETS = {
    "自助餐": "buffet",
    "特色餐": "special",
    "水果": "fruit",
    "冰箱雜貨": "grocery",
    "鍋燒": "nabeyaki",
    "麵食": "noodles",
    "甜點(飲料區)": "dessert",
}
RULE_TYPES = {"upgrade", "addon", "option", "info"}


def enabled_of(value: object) -> bool:
    return str(value or "是").strip() != "否"


def clean_name(value: object) -> str | None:
    if value is None:
        return None
    name = str(value).strip()
    if not name or name.startswith("【範例】"):
        return None
    return name


def price_of(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number <= 0:
        return None
    return number


def deal_of(value: object) -> dict | None:
    if value is None or value == "":
        return None
    text = str(value).strip()
    try:
        qty_part, price_part = text.split("=", 1)
        qty = int(qty_part)
        price = float(price_part)
    except ValueError:
        raise SystemExit(f"優惠欄格式錯誤（應為 數量=價格，例如 3=100）: {text!r}")
    if qty < 2 or price <= 0:
        raise SystemExit(f"優惠欄內容不合理: {text!r}")
    return {"qty": qty, "price": price}


def text_of(value: object) -> str | None:
    if value is None or value == "":
        return None
    text = str(value).strip()
    return text or None


def compile_menu(path: Path = XLSX) -> dict:
    wb = load_workbook(path, data_only=True)
    items: list[dict] = []
    rules: list[dict] = []
    seen: set[tuple[str, str]] = set()
    errors: list[str] = []

    for sheet_name, category in SIMPLE_SHEETS.items():
        ws = wb[sheet_name]
        for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = clean_name(row[0] if row else None)
            if not name:
                continue
            key = (category, name)
            if key in seen:
                errors.append(f"{sheet_name}!A{index} duplicate {name}")
                continue
            price = price_of(row[1] if len(row) > 1 else None)
            if price is None:
                errors.append(f"{sheet_name}!B{index} missing price for {name}")
                continue
            seen.add(key)
            item: dict = {
                "category": category,
                "name": name,
                "price": price,
                "enabled": enabled_of(row[2] if len(row) > 2 else "是"),
            }
            subcategory = text_of(row[3] if len(row) > 3 else None)
            if subcategory:
                item["subcategory"] = subcategory
            deal = deal_of(row[4] if len(row) > 4 else None)
            if deal:
                item["deal"] = deal
            box_price = price_of(row[5] if len(row) > 5 else None)
            if box_price is not None:
                item["boxPrice"] = box_price
            note = text_of(row[6] if len(row) > 6 else None)
            if note:
                item["note"] = note
            items.append(item)

    drink_ws = wb["飲料"]
    for index, row in enumerate(drink_ws.iter_rows(min_row=2, values_only=True), start=2):
        name = clean_name(row[0] if row else None)
        if not name:
            continue
        key = ("drink", name)
        if key in seen:
            errors.append(f"飲料!A{index} duplicate {name}")
            continue
        prices = {}
        for offset, size in enumerate(SIZE_KEYS, start=1):
            amount = price_of(row[offset] if row and len(row) > offset else None)
            if amount is not None:
                prices[size] = amount
        if not prices:
            errors.append(f"飲料!A{index} {name} has no sizes")
            continue
        seen.add(key)
        item: dict = {
            "category": "drink",
            "name": name,
            "prices": prices,
            "enabled": enabled_of(row[5] if len(row) > 5 else "是"),
        }
        subcategory = text_of(row[6] if len(row) > 6 else None)
        if subcategory:
            item["subcategory"] = subcategory
        note = text_of(row[7] if len(row) > 7 else None)
        if note:
            item["note"] = note
        items.append(item)

    rules_ws = wb["加購規則"]
    for index, row in enumerate(rules_ws.iter_rows(min_row=2, values_only=True), start=2):
        rule_type = text_of(row[0] if row else None)
        if rule_type not in RULE_TYPES:
            errors.append(f"加購規則!A{index} unknown rule type {rule_type!r}")
            continue
        rule: dict = {
            "type": rule_type,
            "appliesTo": text_of(row[1] if len(row) > 1 else None) or "",
            "name": text_of(row[2] if len(row) > 2 else None) or "",
        }
        if rule_type in ("upgrade", "addon"):
            add_price = price_of(row[4] if len(row) > 4 else None)
            if add_price is None:
                errors.append(f"加購規則!E{index} missing price for {rule['name']}")
                continue
            rule["price"] = add_price
        content = text_of(row[3] if len(row) > 3 else None)
        if content:
            rule["content"] = content
        note = text_of(row[5] if len(row) > 5 else None)
        if note:
            rule["note"] = note
        rules.append(rule)

    settings = {row[0]: row[1] for row in wb["設定草稿"].iter_rows(min_row=2, values_only=True) if row and row[0]}
    if errors:
        raise SystemExit("menu import failed:\n" + "\n".join(errors))

    return {
        "restaurantName": settings.get("餐廳名稱") or "",
        "mealUnitPriceDefault": float(settings.get("換算_一餐預設") or 100),
        "drinkUnitPriceDefault": float(settings.get("換算_一杯預設") or 30),
        "timezone": "Asia/Taipei",
        "items": items,
        "rules": rules,
    }


def main() -> None:
    menu = compile_menu()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(menu, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    counts: dict[str, int] = {}
    for item in menu["items"]:
        counts[item["category"]] = counts.get(item["category"], 0) + 1
    print(f"wrote {OUT} ({sum(counts.values())} items) {counts}")
    print(f"rules: {len(menu['rules'])}")


if __name__ == "__main__":
    main()
