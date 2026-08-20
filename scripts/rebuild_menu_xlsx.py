"""Rebuild data/menu.xlsx with the new category/subcategory/deal/box/addon schema."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "menu.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="EDEDED")
HEAD_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

SIMPLE_HEADER = ["品項", "價格", "上架", "子分類", "優惠", "整箱價", "備註"]
DRINK_HEADER = ["品項", "熱小", "熱中", "冰中", "特大", "上架", "子分類", "備註"]

def sheet_simple(ws, rows: list[list]) -> None:
    ws.append(SIMPLE_HEADER)
    for row in rows:
        ws.append(row)
    _style(ws, widths=[22, 8, 6, 12, 10, 10, 26])

def sheet_drink(ws, rows: list[list]) -> None:
    ws.append(DRINK_HEADER)
    for row in rows:
        ws.append(row)
    _style(ws, widths=[20, 7, 7, 7, 7, 6, 12, 20])

def _style(ws, widths: list[int]) -> None:
    for cell in ws[1]:
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP

wb = Workbook()

# ── 填寫說明 ──
ws = wb.active
ws.title = "填寫說明"
ws["A1"] = "餐卡菜單填寫表"
ws["A1"].font = Font(bold=True, size=14)
notes = [
    ("目的", "把真實菜單填進來；改完後執行 python scripts/menu_xlsx_to_json.py 重建 src/data/menu.json。"),
    ("分頁", "自助餐／特色餐／水果／冰箱雜貨／鍋燒／麵食／飲料／甜點(飲料區)／加購規則／設定草稿。"),
    ("子分類", "每張表都要填。App 會把同子分類的品項排在一起，例如飲料=咖啡/一般飲料/新品。"),
    ("優惠", "品項單價照常填；有量販優惠時填「數量=價格」，例如蘋果一顆 35、優惠欄填 3=100。App 會自動算 4 顆=100+35=135。"),
    ("整箱價", "冰箱雜貨專用。瓶價填「價格」，箱價填「整箱價」。App 預設只顯示瓶裝，需要時才點「買一箱」。"),
    ("加購規則", "麵食的套餐升級、加料價、麵體選擇都記在「加購規論」分頁；櫃檯規則變了只改那頁。"),
    ("上架", "填「是」或「否」；空白視為「是」。暫停販售填否即可，不必刪列。"),
    ("名稱", "同一分頁內不要重複，名稱跟櫃檯一致方便對帳。"),
]
for row in notes:
    ws.append(list(row))
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 90
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = WRAP

# ── 自助餐 ──
ws = wb.create_sheet("自助餐")
sheet_simple(ws, [
    ["基本餐", 70, "是", "主食", None, None, "基本主菜+3副菜"],
    ["換購A餐", 95, "是", "主食", None, None, "A主菜+3副菜"],
    ["換購B餐", 115, "是", "主食", None, None, "B主菜+3副菜"],
    ["加購A餐", 120, "是", "主食", None, None, "基本主菜+A主菜+3副菜"],
    ["加購B餐", 140, "是", "主食", None, None, "基本主菜+B主菜+3副菜"],
    ["單點A", 60, "是", "主菜", None, None, "A主菜"],
    ["單點B", 80, "是", "主菜", None, None, "B主菜"],
    ["外帶盒", 5, "是", "加購", None, None, "加購"],
])

# ── 特色餐 ──
ws = wb.create_sheet("特色餐")
sheet_simple(ws, [
    ["健康餐-蒜香肉片", 120, "是", "健康餐", None, None, None],
    ["健康餐-迷迭香雞片", 120, "是", "健康餐", None, None, None],
    ["特色餐-香酥大雞腿", 110, "是", "特色餐", None, None, None],
])

# ── 水果 ──
ws = wb.create_sheet("水果")
sheet_simple(ws, [
    ["蘋果", 35, "是", "水果", "3=100", None, None],
    ["柳丁/柑橘", 20, "是", "水果", "6=110", None, None],
    ["藍莓", 80, "是", "水果", None, None, None],
    ["酪梨", 55, "是", "水果", None, None, None],
    ["水梨", 45, "是", "水果", None, None, None],
    ["黃金奇異果", 25, "是", "水果", None, None, None],
    ["紅色火龍果", 60, "是", "水果", None, None, None],
])

# ── 冰箱雜貨 ──
ws = wb.create_sheet("冰箱雜貨")
sheet_simple(ws, [
    ["寶佳麗強氣泡水", 30, "是", "寶佳麗", None, 720, "330mL；1箱=24瓶"],
    ["寶佳麗通寧汽水", 50, "是", "寶佳麗", None, 1200, "330mL；1箱=24瓶"],
    ["職人動物福利鮮奶", 80, "是", "鮮奶", None, None, "936mL"],
    ["愛之味燕麥奶", 120, "是", "愛之味", None, 2880, "990mL；1箱=24瓶"],
    ["波蜜元氣海礦補給飲", 30, "是", "波蜜", None, 720, "600mL；1箱=24瓶"],
    ["養樂多豆漿", 15, "是", "養樂多", None, 360, "200mL；1箱=24瓶"],
    ["養樂多蘋果汁", 15, "是", "養樂多", None, 360, "200mL；1箱=24瓶"],
    ["養樂多超無限運動飲料", 20, "是", "養樂多", None, 480, "320mL；1箱=24瓶"],
    ["養樂多優格碳酸飲料", 75, "是", "養樂多", None, 1800, "300mL；1箱=24瓶"],
    ["愛之味莎莎亞椰奶", 20, "是", "愛之味", None, 480, "250mL；1箱=24瓶"],
    ["波蜜紫色蔬果汁", 20, "是", "波蜜", None, 360, "250mL；1箱=18瓶"],
    ["咖啡師燕麥奶", 120, "是", "燕麥奶", None, None, None],
])

# ── 鍋燒 ──
ws = wb.create_sheet("鍋燒")
sheet_simple(ws, [
    ["原味鍋燒", 80, "是", "鍋燒", None, None, None],
    ["沙茶鍋燒", 85, "是", "鍋燒", None, None, None],
    ["韓式鍋燒", 85, "是", "鍋燒", None, None, None],
    ["加蛋", 10, "是", "加購", None, None, None],
    ["加麵", 10, "是", "加購", None, None, None],
])

# ── 麵食 ──
ws = wb.create_sheet("麵食")
sheet_simple(ws, [
    ["古早味乾麵", 63, "是", "麵食水餃", None, None, "基本餐；597卡；外包68"],
    ["單純風味乾拌麵", 50, "是", "麵食水餃", None, None, "麻醬/麻辣/肉臊；507卡"],
    ["素拌麵", 50, "是", "麵食水餃", None, None, "527卡"],
    ["乾拌麵套餐", 80, "是", "麵食水餃", None, None, "686卡"],
    ["素拌麵套餐", 80, "是", "麵食水餃", None, None, "718卡"],
    ["餛飩麵", 80, "是", "麵食水餃", None, None, "533卡"],
    ["沙茶乾麵", 80, "是", "麵食水餃", None, None, "530卡"],
    ["榨菜肉絲麵", 80, "是", "麵食水餃", None, None, "589卡"],
    ["銷魂皮蛋肉醬麵", 80, "是", "麵食水餃", None, None, "634卡"],
    ["一份水餃", 84, "是", "麵食水餃", None, None, "625卡"],
    ["半份水餃", 45, "是", "麵食水餃", None, None, None],
    ["燙青菜", 40, "是", "麵食水餃", None, None, "98卡"],
    ["清蒸肉圓+蘿蔔貢丸湯", 75, "是", "台灣小吃", None, None, "500卡"],
    ["米糕+蘿蔔貢丸湯", 75, "是", "台灣小吃", None, None, "615卡"],
    ["鹿港芋丸+蘿蔔貢丸湯", 75, "是", "台灣小吃", None, None, "600卡"],
    ["蘿蔔貢丸湯", 40, "是", "台灣小吃", None, None, "118卡"],
    ["淡水魚丸湯", 40, "是", "台灣小吃", None, None, "115卡"],
    ["餛飩湯", 40, "是", "台灣小吃", None, None, "120卡"],
])

# ── 飲料 ──
ws = wb.create_sheet("飲料")
sheet_drink(ws, [
    ["每日咖啡", 40, 45, 40, 65, "是", "咖啡", None],
    ["美式咖啡", 45, 55, 45, 75, "是", "咖啡", None],
    ["特調咖啡", 45, 55, 45, 80, "是", "咖啡", None],
    ["拿鐵咖啡", 55, 65, 55, 95, "是", "咖啡", None],
    ["榛果拿鐵", 65, 80, 65, 110, "是", "咖啡", None],
    ["香草拿鐵", 65, 80, 65, 110, "是", "咖啡", None],
    ["焦糖拿鐵", 65, 80, 65, 110, "是", "咖啡", None],
    ["紅茶", 20, 25, 25, 40, "是", "一般飲料", None],
    ["奶茶", 25, 30, 30, 45, "是", "一般飲料", None],
    ["風味奶茶", 35, 40, 40, 55, "是", "一般飲料", None],
    ["紅茶拿鐵", 40, 45, 45, 65, "是", "一般飲料", None],
    ["茉莉綠茶", 30, 35, 35, 50, "是", "一般飲料", None],
    ["綠奶茶", 35, 40, 40, 55, "是", "一般飲料", None],
    ["風味綠奶茶", 45, 50, 50, 65, "是", "一般飲料", None],
    ["綠茶拿鐵", 50, 55, 55, 75, "是", "一般飲料", None],
    ["蜂蜜綠茶", None, None, 45, 60, "是", "一般飲料", None],
    ["百香綠(紅)茶", None, 65, 65, 80, "是", "一般飲料", None],
    ["香橙冰茶", None, None, 50, 70, "是", "一般飲料", None],
    ["經典可可", 55, 60, 60, 70, "是", "一般飲料", None],
    ["可可拿鐵", 75, 80, 80, 105, "是", "一般飲料", None],
    ["風味可可", 65, 70, 70, 90, "是", "一般飲料", None],
    ["抹茶拿鐵", None, 75, 75, 95, "是", "一般飲料", None],
    ["蜂蜜柚子茶", None, 75, 75, 90, "是", "一般飲料", None],
    ["現打果汁牛奶", None, None, 75, 95, "是", "一般飲料", None],
    ["珍珠鮮奶茶", None, None, None, 75, "是", "一般飲料", None],
    ["全脂牛奶", None, None, 51, 70, "是", "一般飲料", None],
    ["蜂蜜牛奶", None, None, 60, 80, "是", "一般飲料", None],
    ["可可咖啡", None, None, None, 90, "是", "新品", None],
    ["柚香咖啡", None, None, None, 80, "是", "新品", None],
    ["黑糖咖啡", None, None, None, 80, "是", "新品", None],
    ["蜂蜜咖啡", None, None, None, 80, "是", "新品", None],
    ["泰式奶茶", None, None, None, 65, "是", "新品", None],
])

# ── 甜點(飲料區) ──
ws = wb.create_sheet("甜點(飲料區)")
sheet_simple(ws, [
    ["熱壓吐司", 55, "是", "甜點", None, None, None],
])

# ── 加購規則 ──
ws = wb.create_sheet("加購規則")
ws.append(["類型", "套用對象", "名稱", "內容", "加價", "備註"])
rules = [
    ["upgrade", "麵食:麵食水餃", "套餐A", "豆干2片+滷蛋1個+豆芽增量", 30, None],
    ["upgrade", "麵食:麵食水餃", "套餐B", "肉片5片+淡水魚丸1個", 30, None],
    ["upgrade", "麵食:麵食水餃", "套餐C", "豆干2片+素肉燥+豆芽增量", 30, "只限素拌麵"],
    ["addon", "麵食:麵食水餃", "肉片", "2片", 10, None],
    ["addon", "麵食:麵食水餃", "豆干", "2片", 10, None],
    ["addon", "麵食:麵食水餃", "小貢丸", "2個", 10, None],
    ["addon", "麵食:麵食水餃", "淡水魚丸", "2個", 10, None],
    ["addon", "麵食:麵食水餃", "餛飩", "2個", 10, None],
    ["addon", "麵食:麵食水餃", "加麵", "1份", 10, None],
    ["addon", "麵食:麵食水餃", "滷蛋", "1個", 15, None],
    ["addon", "鍋燒:加購", "加蛋", "1個", 10, None],
    ["addon", "鍋燒:加購", "加麵", "1份", 10, None],
    ["option", "麵食:麵食水餃", "麵體", "油麵/白麵/意麵/米粉/粄條麵", 0, "自由選擇，不加價"],
    ["info", "麵食:麵食水餃", "供應時段", "週一至週五 11:00~18:30", None, None],
]
for row in rules:
    ws.append(row)
_style(ws, widths=[10, 18, 14, 40, 8, 18])

# ── 設定草稿 ──
ws = wb.create_sheet("設定草稿")
ws.append(["鍵", "值", "說明"])
for row in [
    ["餐廳名稱", "員工餐廳", "選填，之後顯示在 App 標題"],
    ["換算_一餐預設", 80, "之後可在 App 改"],
    ["換算_一杯預設", 60, "之後可在 App 改"],
    ["時區", "Asia/Taipei", "月底歸零、日均都用這個時區"],
    ["結算", "每月1號到月底", "餘額月底丟掉，不要改這格除非規則變了"],
]:
    ws.append(row)
_style(ws, widths=[16, 22, 44])

wb.save(XLSX)
print(f"rewrote {XLSX}")
print("sheets:", wb.sheetnames)
